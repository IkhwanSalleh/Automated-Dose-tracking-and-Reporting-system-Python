import pandas
from openpyxl import load_workbook
import time
import random
import Studygrouplibrary as Lib2 

def append_df_to_excel(filename, df, sheet_name,startcol, startrow=None ,truncate_sheet=False,**to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]

    Returns: None
    """
    writer = pandas.ExcelWriter(filename, engine='openpyxl')
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)

        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row

        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)

        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name,startcol, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
    
    
def sheet_selections_Study(results):
    StudyGroup,StudyCode=Lib2.typeofscanstudy(results)
    Compileddataexcel=results
    condition_not_met=0
    while condition_not_met <=100: 
        try:    
            if StudyCode ==1:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT ABDOMEN',startcol=0, header=None)
            elif StudyCode ==2:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT CARDIAC',startcol=0, header=None)
            elif StudyCode ==3:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT BRAIN',startcol=0, header=None)
            elif StudyCode ==4:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT CHEST',startcol=0, header=None)
            elif StudyCode ==5:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT HEAD',startcol=0, header=None)
            elif StudyCode ==6:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT PELVIS',startcol=0, header=None)
            elif StudyCode ==7:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT SPINE',startcol=0, header=None)
            elif StudyCode ==8:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT THORAX',startcol=0, header=None)
            elif StudyCode ==9:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT WHOLE BODY',startcol=0, header=None)     
            else:
                append_df_to_excel('DATAFOLEDER/DATAstudy.xlsx',Compileddataexcel,sheet_name='CT OTHER',startcol=0, header=None)
            break
        except:
            secosleep=random.randint(1,5)
            time.sleep(secosleep)
            print('fail study')
            print(StudyCode)
                
def Sheet_selection_series_UPPER(results):
    
    scancounter=0
    condition_not_met=0
    Typeofscan=results['Scan']
    DLP=results['DLP']
    CTDIvol=results['CTDIvol']
    PatientID=results['PatientID']
    
    for scanlist in Typeofscan:
        Compiledseries=pandas.DataFrame({'PatientID':PatientID[0],'Scan':Typeofscan[scancounter],'DLP':DLP[scancounter],'CTDIvol':CTDIvol[scancounter],}, index=[0])
        if Typeofscan is None:
            break
        elif 'BRAIN' in str(scanlist):
            while condition_not_met <=100:
                if 'PLAIN' in str(scanlist):
                    try:
                        append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-BRAIN',startcol=0, header=None)
                        break
                    except:
                        secosleep=random.randint(1,5)
                        time.sleep(secosleep)
                        print('fail head-brain')
                elif 'ORBIT' in str(scanlist):
                    try:
                        append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-BRAIN',startcol=4, header=None)
                        break
                    except:
                        secosleep=random.randint(1,5)
                        time.sleep(secosleep)
                        print('fail head-brain')
                elif 'SPIRAL' in str(scanlist):
                    try:
                        append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-BRAIN',startcol=9, header=None)
                        break
                    except:
                        secosleep=random.randint(1,5)
                        time.sleep(secosleep)
                        print('fail head-brain')
                elif 'CTA' in str(scanlist):
                    try:
                        append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-BRAIN',startcol=14, header=None)
                        break
                    except:
                        secosleep=random.randint(1,5)
                        time.sleep(secosleep)
                        print('fail head-brain')
                elif 'IV' in str(scanlist):
                    try:
                        append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-BRAIN',startcol=19, header=None)
                        break
                    except:
                        secosleep=random.randint(1,5)
                        time.sleep(secosleep)
                        print('fail head-brain')
                else:
                    try:
                        append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-BRAIN',startcol=15, header=None)
                        break
                    except:
                        secosleep=random.randint(1,5)
                        time.sleep(secosleep)
                        print('fail head-brain')
                    
            
        elif 'HEAD' in str(scanlist):
            while condition_not_met <=100:
                 try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-HEAD',startcol=0, header=None)
                    break
                 except:
                     secosleep=random.randint(1,5)
                     time.sleep(secosleep)
                     print('fail head-Head')
                     
        elif 'NECK' in str(scanlist):
            while condition_not_met <=100:
                 try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-NECK',startcol=0, header=None)
                    break
                 except:
                     secosleep=random.randint(1,5)
                     time.sleep(secosleep)
                     print('fail head-Neck')
                     
        elif 'SINUS' in str(scanlist):
            while condition_not_met <=100:
                 try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-SINUS',startcol=0, header=None)
                    break
                 except:
                     secosleep=random.randint(1,5)
                     time.sleep(secosleep)
                     print('fail head-Neck') 
                     
        elif 'EAR' in str(scanlist):
            while condition_not_met <=100:
                 try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-EAR',startcol=0, header=None)
                    break
                 except:
                     secosleep=random.randint(1,5)
                     time.sleep(secosleep)
                     print('fail head-Neck')  
        else:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyUpper.xlsx',Compiledseries,sheet_name='CT HEAD-Other',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail head ')
                    
            
        scancounter=scancounter+1
        
        
def Sheet_selection_series_MIDDLE(results):
    scancounter=0
    condition_not_met=0
    Typeofscan=results['Scan']
    DLP=results['DLP']
    CTDIvol=results['CTDIvol']
    PatientID=results['PatientID']
    scancounter=0
    condition_not_met=0
    for scanlist in Typeofscan:
        Compiledseries=pandas.DataFrame({'Scan':Typeofscan[scancounter],'DLP':DLP[scancounter],'CTDIvol':CTDIvol[scancounter],'PatientID':PatientID[0]}, index=[0])
        if Typeofscan is None:
            break
        elif 'ABDOMEN,PELVIS ' in str(scanlist):
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Abdomen,pelvis',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
        elif 'LIVER' in scanlist:
            while condition_not_met <=100:
                 try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Liver',startcol=0, header=None)
                    break
                 except:
                     secosleep=random.randint(1,5)
                     time.sleep(secosleep)
                     print('fail abdomen')
                     
            
        elif 'ARTERIAL' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Arterial Phase',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
        elif 'VENOUS' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Venous Phase',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
        elif 'DELAYED' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Delayed Phase',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                
        elif 'ABDOMEN' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Abdomen',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                   
                    
                    
        elif 'CAP' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-CAP',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                
        elif 'CTA' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-CTA',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                   
                
        elif 'CTU' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-CTU',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                    
        elif 'BLADDER' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Bladder',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                   
                    
        elif 'FLUORO' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Fluoro',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                    
        elif 'KIDNEY' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Kidney',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                    
        elif 'LIVER,PANCREASE' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Liver,Pancrease',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                    
        elif 'PELVIS' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Pelvis',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                    
        elif 'PLAIN' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Plain',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
                
        elif 'RENAL' in scanlist:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Renal',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
        else:
            while condition_not_met <=100:
                try:
                    append_df_to_excel('DATAFOLEDER/DATAstudyMiddle.xlsx',Compiledseries,sheet_name='CT ABDOMEN-Other',startcol=0, header=None)
                    break
                except:
                    secosleep=random.randint(1,5)
                    time.sleep(secosleep)
                    print('fail abdomen')
                    
    
    
            
        scancounter=scancounter+1